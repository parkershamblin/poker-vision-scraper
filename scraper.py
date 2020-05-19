# Input the full file path to the chrome driver v81 exe you downloaded here.
chromedriver_path = r"ENTER THE FULL FILE PATH TO 'chromedriver.exe' YOU DOWNLOADED HERE"
# Input your Run It Once Log In Details here.
username_input = "ENTER YOUR RUN IT ONCE USERNAME HERE"
password_input = "ENTER YOUR RUN IT ONCE PASSWORD HERE"

import time
import win32api
from os import path
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Load workbook if it exist and if not create new workbook.
if path.exists('RIO-Vision-Results.xlsx'):
    wb = load_workbook('RIO-Vision-Results.xlsx')
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = 'RIO-Vision-Results'
    ws['A1'] = 'Line'
    ws['B1'] = 'Overview'
    ws['C1'] = 'Board'
    ws['D1'] = 'Hand'
    ws['E1'] = 'Result'



#create webdriver object and get url
driver = webdriver.Chrome(chromedriver_path)
driver.implicitly_wait(1)
driver.get('https://www.runitonce.com/login/?next=/vision/')


# log into Run It Once
time.sleep(7)  # wait for login page to load
username = driver.find_element_by_id('login-username')
username.clear()
username.send_keys(username_input)
password = driver.find_element_by_id('login-password')
password.clear()
password.send_keys(password_input)
driver.find_element_by_class_name('login-btn').click()


previous_result = ""
previous_hand = ""


def reverse(s): 
  string = "" 
  for i in s:
    string = i + string
  return string


def scrape_vision():
    try:
        line = driver.find_element_by_id('board-header').text
        board = driver.find_element_by_class_name('board').get_attribute('data-texture')
        intro = driver.find_element_by_id('intro-text')

        hand = ""
        for i in intro.find_elements_by_class_name('result-hand-graphical'):
            hand += reverse(i.get_attribute('src').replace(r"https://www.runitonce.com/static/img/visions/card-set/", "").replace(r".png?2.0", "").replace('/', ''))

        # grab result and print text
        result = driver.find_element_by_id('practice-result')
        if len(result.text) > 1:
            global previous_result
            global previous_hand
            if previous_result != result.text and previous_hand != hand:
                    html = driver.page_source
                    parsed_html = BeautifulSoup(html, 'html.parser')
                    situation = parsed_html.body.find('h5', {'id': 'hand-graph-title-overview'}).text
                    if len(situation) > 1:
                        ws.append([line, situation, board, hand, result.text])
                        wb.save('RIO-Vision-Results.xlsx')
                        previous_result = result
                        previous_hand = hand
                        print(situation, result.text, hand, sep="\n")
    except Exception as e:
        print(f"Error: {e}")
        time.sleep(1)
        pass


def check_for_left_click():
    state_left = win32api.GetKeyState(0x01)
    while True:
        a = win32api.GetKeyState(0x01)
        if a != state_left:
            state_left = a
            if a < 0:
                time.sleep(0.05)
                scrape_vision()
        time.sleep(0.001)


check_for_left_click()